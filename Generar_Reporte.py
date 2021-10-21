import datetime
import pandas as pd
import openpyxl as xl
import shutil
import time


def copyData(name,data):
	"""
		Completa los campos del documento de excel
	"""
	srcfile = xl.load_workbook(name)
	sheet 	= srcfile['Hoja1']
	unique = data.groupby('Reporte').nunique().drop(['Responsable','Reportado_por', 'Fecha'], axis=1).rename(columns={"Canaleta": "Frecuencia"})
	
	for row in range(4, len(data) + 4):
		sheet.cell(row=row,column=1).value = data['Canaleta'][row -4]
		sheet.cell(row=row,column=2).value = data['Reporte'][row -4]
		sheet.cell(row=row,column=3).value = data['Responsable'][row -4]
		sheet.cell(row=row,column=4).value = data['Reportado_por'][row -4]
		sheet.cell(row=row,column=5).value = data['Fecha'][row -4]

	uniques = list(unique.index.values)
	valuesU = list(unique['Frecuencia'])
	for row in range(len(uniques)):
		sheet.cell(row=row+4,column=7).value = uniques[row]
		sheet.cell(row=row+4,column=8).value = valuesU[row]
	
	srcfile.save(name)

def createNewReport(name):
	"""
		Hace copia de la plantilla_reporte
	"""
	original = 'plantilla_reporte.xlsx'
	target = f'{name}'
	shutil.copyfile(original, target)

def getFilteredData(start, end):
	"""
		Lee datos desde csv y devuelve los datos filtrados por fechas
	"""
	try:
		data = pd.read_csv('reports.csv')
		data = data[data['Fecha'].between(start, end)]
		data['Fecha'] = data['Fecha'].apply(getDate)
		return data
	except Exception as e:
		print("Error getFilteredData, ", e)

def getDate(date):
	"""
		Retorna la fecha en texto
	"""
	return datetime.datetime.fromtimestamp(date).strftime('%d-%m-%Y %H:%M:%S')

def getTimestamp(date):
	"""
		Devuelve la fecha en unix epoch
	"""
	return time.mktime(datetime.datetime.strptime(date, '%d-%m-%Y %H:%M:%S').timetuple())

def requestRange(start, end):
	"""
		convierte rango de fechas de entrada a unix epoch
	"""
	try:
		startDate  		= start + ' 00:00:00'
		endDate  		= end + ' 00:00:00'
		startTimestamp 	= getTimestamp(startDate)
		endTimestamp 	= getTimestamp(endDate)
		return startTimestamp, endTimestamp
	except Exception as e:
		print('Ingresa las fechas con el formato correcto', e)

